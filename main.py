#this project is to analyze the performance of Arsenal players in both match Arsenal vs Leicester and Arsenal vs leeds
#the results are the best performance of players in order for each match
#the best cliques in the team
#most frequent passes during each match


import openpyxl as xl
import networkx as nx
from most_important_player import playersImportance
from most_possible_passes import mostPossiblePasses
from matplotlib import pyplot as plt
import find_good_cliques as fgc
from math import log10, floor



def round_sig(x, sig=2):
    if(x != 0):
        return round(x, sig-int(floor(log10(abs(x))))-1)
    else:
        return 0

#first match:

match_1 = xl.load_workbook("Book1.xlsx")
sheet_match1 = match_1["Sheet1"]
names_column_match1 = 3
positions_column_match1 = 2
players_numbers_column_match1 = 1
match1 = nx.DiGraph()


#second match:

match_2 = xl.load_workbook("2-Arsenal vs Leicester (13 Mar 2022).xlsx")
sheet_match2 = match_2["Sheet1"]
names_column_match2 = 3
positions_column_match2 = 2
players_numbers_column_match2 = 1
match2 = nx.DiGraph()


def generateGraph(position_x, position_y, number_of_players, sheet, match):
    #the position x and y are the row and column of the name in the xl_sheet
    graph = nx.DiGraph()
    for number in range(1, number_of_players+1):
        graph.add_node(number)

    for row in range(position_x + 1, position_x + 1 + number_of_players):
        graph._node[row-position_x]['name'] = sheet.cell(row, names_column_match1).value
        graph._node[row-position_x]['position'] = sheet.cell(row, positions_column_match1).value
        graph._node[row-position_x]['number'] = sheet.cell(row, players_numbers_column_match1).value

    for row in range(1, number_of_players+1):
        for column in range(1, number_of_players+1):
            if sheet.cell(row + position_x, column + position_y).value > 0:
                graph.add_edge(row, column, weight=sheet.cell(row + position_x, column + position_y).value)
                if (row, column) in match.edges():
                    list(match.edges(data=True))[list(match.edges()).index((row, column))][2]['weight'] +=\
                        sheet.cell(row + position_x, column + position_y).value
                else:
                    match.add_edge(row, column, weight=sheet.cell(row + position_x, column + position_y).value)
    return (graph, match)


#first match loading

graph_0_10_1, match1 = generateGraph(5, 3, 14, sheet_match1, match1)
graph_10_20_1, match1 = generateGraph(23, 3, 14, sheet_match1, match1)
graph_20_30_1, match1 = generateGraph(41, 3, 14, sheet_match1, match1)
graph_30_40_1, match1 = generateGraph(59, 3, 14, sheet_match1, match1)
graph_40_50_1, match1 = generateGraph(77, 3, 14, sheet_match1, match1)
graph_50_60_1, match1 = generateGraph(95, 3, 14, sheet_match1, match1)
graph_60_70_1, match1 = generateGraph(113, 3, 14, sheet_match1, match1)
graph_70_80_1, match1 = generateGraph(131, 3, 14, sheet_match1, match1)
graph_80_90_1, match1 = generateGraph(149, 3, 14, sheet_match1, match1)

graphs_match1 = [graph_0_10_1, graph_10_20_1, graph_20_30_1, graph_30_40_1, graph_40_50_1, graph_50_60_1, graph_60_70_1, graph_70_80_1, graph_80_90_1]

#second match loading

graph_0_10_2, match2 = generateGraph(5, 3, 14, sheet_match2 , match2)
graph_10_20_2, match2 = generateGraph(23, 3, 14, sheet_match2, match2)
graph_20_30_2, match2 = generateGraph(41, 3, 14, sheet_match2, match2)
graph_30_40_2, match2 = generateGraph(59, 3, 14, sheet_match2, match2)
graph_40_50_2, match2 = generateGraph(77, 3, 14, sheet_match2, match2)
graph_50_60_2, match2 = generateGraph(95, 3, 14, sheet_match2, match2)
graph_60_70_2, match2 = generateGraph(113, 3, 14, sheet_match2, match2)
graph_70_80_2, match2 = generateGraph(131, 3, 14, sheet_match2, match2)
graph_80_90_2, match2 = generateGraph(149, 3, 14, sheet_match2, match2)

graphs_match2 = [graph_0_10_2, graph_10_20_2, graph_20_30_2, graph_30_40_2, graph_40_50_2, graph_50_60_2, graph_60_70_2, graph_70_80_2, graph_80_90_2]


for graph in graphs_match1:
    for node in graph.nodes(data=True):
        if not (node in list(match1.nodes(data=True))):
            match1.add_node(node[0])
            match1._node[node[0]]['name'] = node[1]['name']
            match1._node[node[0]]['position'] = node[1]['position']
            match1._node[node[0]]['number'] = node[1]['number']


for node in graphs_match2[0].nodes(data=True): #any graph will work and node is a tuplw
    match2.add_node(node[0])
    match2._node[node[0]]['name'] = node[1]['name']
    match2._node[node[0]]['position'] = node[1]['position']
    match2._node[node[0]]['number'] = node[1]['number']


def getInitials(name):
    names = name.split(' ')
    Initials = names[0][0]
    for i in range(1, len(names)):
        Initials += ('.' + names[i][0])

    return Initials

def printClustering(graph):
    cluster = nx.cluster.clustering(graph)
    output = ''
    for k, v in cluster.items():
        output += f'{getInitials(dict(graph.nodes(data=True))[k]["name"])}: {round_sig(v)}, '

    print('Clustering: ' + output)

def printPageRank(graph):
    pageRank = nx.pagerank(graph)
    output = ''
    for k, v in pageRank.items():
        output += f'{getInitials(dict(graph.nodes(data=True))[k]["name"])}: {round_sig(v)}, '

    print("Page Rank: " + output)


def printBetweenness(graph):
    betweenness = nx.betweenness_centrality(graph)
    output = ''
    for k, v in betweenness.items():
        output += f'{getInitials(dict(graph.nodes(data=True))[k]["name"])}: {round_sig(v)}, '

    print("Betweenness: " + output)


def printResults(graph, graphs_list, name):
    print(name + '\n\n')
    cliques = list(nx.enumerate_all_cliques(match1.to_undirected()))
    print("Best Cliques in the team:\n")
    # to find all good cliques
    best_cliques = []  # if each node has an out degree higher than 10 then its a good clique

    for clq in cliques:
        if fgc.isGoodClique(clq, graph, 10) and len(list(clq)) == 3:
            best_cliques.append(clq)

    for clq in best_cliques:
        print("clique: " + str(best_cliques.index(clq)))
        fgc.printClq(clq, match1, "name")
        print("\n\n")

    print("Players performance:\n")
    sorted_players = playersImportance(graph, "name")
    for k, v in sorted_players.items():
        print(f"{k}: {v}")

    print("\n\n")

    print("Most possible passes:\n")
    most_possible_passes = mostPossiblePasses(graph, "name")
    for Pass in most_possible_passes:
        print(Pass)

    print('\n\n')

    print('stats:')
    for i in range(0, len(graphs_list)):
        printClustering(graphs_list[i])
        printBetweenness(graphs_list[i])
        printPageRank(graphs_list[i])


printResults(match1, graphs_match1, "match 1")

printResults(match2, graphs_match2, "match 2")

match1_label_dict = {}
match1_nodes = list(match1.nodes(data=True))

match2_label_dict = {}
match2_nodes = list(match2.nodes(data=True))

for i in range(14):
    match1_label_dict[match1_nodes[i][0]] = match1_nodes[i][1]['name']
    match2_label_dict[match2_nodes[i][0]] = match2_nodes[i][1]['name']


nx.draw_circular(match1, labels=match1_label_dict, with_labels=True)
plt.show()


''''
nx.draw_circular(match2, labels=match2_label_dict, with_labels=True)
plt.show()
'''
